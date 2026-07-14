from __future__ import annotations

import re
import tempfile
from dataclasses import replace
from html.parser import HTMLParser
from io import BytesIO
from pathlib import Path
from typing import Any
from zipfile import BadZipFile, ZipFile

from ..schemas import MaterialCapability, OutlineItem, DocumentLink, ParsedDocument, ParseWarning, SourceFile, SpreadsheetCell, SpreadsheetSheet, SpreadsheetWorkbook, StyleHint

try:  # pragma: no cover - optional dependency
    from markitdown import MarkItDown  # type: ignore
except ModuleNotFoundError:  # pragma: no cover - import guard
    MarkItDown = None

try:  # pragma: no cover - optional dependency
    from openpyxl import load_workbook  # type: ignore
except ModuleNotFoundError:  # pragma: no cover - import guard
    load_workbook = None


class _TextExtractor(HTMLParser):
    def __init__(self) -> None:
        super().__init__()
        self.chunks: list[str] = []
        self.headings: list[tuple[int, str]] = []
        self.links: list[DocumentLink] = []
        self._heading_level: int | None = None
        self._heading_buffer: list[str] = []
        self._anchor_href: str = ""
        self._anchor_buffer: list[str] = []

    def handle_starttag(self, tag: str, attrs: list[tuple[str, str | None]]) -> None:
        attrs_map = {key: value or "" for key, value in attrs}
        if tag in {"h1", "h2", "h3", "h4", "h5", "h6"}:
            self._heading_level = int(tag[1])
            self._heading_buffer = []
        if tag == "a":
            self._anchor_href = attrs_map.get("href", "")
            self._anchor_buffer = []

    def handle_endtag(self, tag: str) -> None:
        if tag in {"h1", "h2", "h3", "h4", "h5", "h6"} and self._heading_level is not None:
            title = normalize_text("".join(self._heading_buffer))
            if title:
                self.headings.append((self._heading_level, title))
            self._heading_level = None
            self._heading_buffer = []
        if tag == "a" and self._anchor_href:
            text = normalize_text("".join(self._anchor_buffer))
            if text:
                self.links.append(DocumentLink(text=text, url=self._anchor_href, source="html"))
            self._anchor_href = ""
            self._anchor_buffer = []

    def handle_data(self, data: str) -> None:
        if self._heading_level is not None:
            self._heading_buffer.append(data)
        if self._anchor_href:
            self._anchor_buffer.append(data)
        self.chunks.append(data)


BASIC_SUFFIXES = {".html", ".htm", ".md", ".markdown", ".txt", ""}
MARKITDOWN_SUFFIXES = {
    ".atom",
    ".csv",
    ".doc",
    ".docx",
    ".epub",
    ".ipynb",
    ".jpeg",
    ".jpg",
    ".json",
    ".jsonl",
    ".pdf",
    ".png",
    ".ppt",
    ".pptx",
    ".rss",
    ".xls",
    ".xlsx",
    ".xml",
    ".zip",
}
PARSER_MARKITDOWN = "markitdown"
PARSER_BASIC = "basic"
PARSER_RETRY_ATTEMPTS = 2
MAX_WORKBOOK_SHEETS = 200
MAX_WORKBOOK_CELLS = 250_000
MAX_XLSX_ARCHIVE_ENTRIES = 20_000
MAX_XLSX_UNCOMPRESSED_BYTES = 1024 * 1024 * 1024
MAX_XLSX_COMPRESSION_RATIO = 250


def parse_document(
    *,
    filename: str,
    content: bytes,
    content_type: str = "",
    reference_role: str = "material",
    parser_mode: str = PARSER_MARKITDOWN,
) -> ParsedDocument:
    suffix = Path(filename).suffix.lower()
    if suffix in BASIC_SUFFIXES or any(token in content_type for token in ("html", "markdown", "plain")):
        return parse_document_basic(filename=filename, content=content, content_type=content_type, reference_role=reference_role)

    if suffix not in MARKITDOWN_SUFFIXES:
        return parse_document_basic(filename=filename, content=content, content_type=content_type, reference_role=reference_role)

    if normalize_parser_mode(parser_mode) == PARSER_BASIC:
        fallback = parse_with_retry(
            lambda: parse_document_basic(filename=filename, content=content, content_type=content_type, reference_role=reference_role),
            parser_name="basic",
        )
        warning = ParseWarning(
            code="enhanced_parser_disabled",
            message="Enhanced document parsing is disabled; used basic parser fallback.",
            severity="info",
        )
        return replace(fallback, warnings=[*fallback.warnings, warning])

    parsed = parse_with_specialized_parser(filename=filename, content=content, content_type=content_type, reference_role=reference_role)
    if parsed is not None and (parsed.plain_text or any(warning.code == "xlsx_resource_boundary" for warning in parsed.warnings)):
        return with_source_file(parsed, filename=filename, content_type=content_type, size=len(content), reference_role=reference_role)

    parsed = parse_with_markitdown(filename, content)
    if parsed is not None and parsed.plain_text:
        return with_source_file(parsed, filename=filename, content_type=content_type, size=len(content), reference_role=reference_role)

    fallback = parse_with_retry(
        lambda: parse_document_basic(filename=filename, content=content, content_type=content_type, reference_role=reference_role),
        parser_name="basic",
    )
    warning = ParseWarning(
        code="markitdown_unavailable_or_failed",
        message="MarkItDown was unavailable or could not extract text; used basic parser fallback.",
        severity="warning",
    )
    if parsed is not None:
        return replace(fallback, warnings=[*parsed.warnings, *fallback.warnings, warning])
    return replace(fallback, warnings=[*fallback.warnings, warning])


def parse_with_specialized_parser(*, filename: str, content: bytes, content_type: str, reference_role: str) -> ParsedDocument | None:
    """Run safe, format-specific parsers before the generic MarkItDown layer."""
    if Path(filename).suffix.lower() == ".xlsx":
        return parse_xlsx_document(filename=filename, content=content)
    return None


def parse_xlsx_document(*, filename: str, content: bytes) -> ParsedDocument | None:
    if load_workbook is None:
        return None
    boundary_reason = xlsx_archive_boundary_reason(content)
    if boundary_reason:
        return ParsedDocument(warnings=[ParseWarning(code="xlsx_resource_boundary", message=boundary_reason, severity="error")])
    try:
        formula_book = load_workbook(BytesIO(content), data_only=False, read_only=False, keep_links=True)
        value_book = load_workbook(BytesIO(content), data_only=True, read_only=False, keep_links=True)
    except Exception:
        return None

    sheets: list[SpreadsheetSheet] = []
    text_parts: list[str] = []
    warnings: list[ParseWarning] = []
    total_cells = 0
    formula_count = 0
    truncated = False
    for sheet_index, sheet in enumerate(formula_book.worksheets):
        if sheet_index >= MAX_WORKBOOK_SHEETS or total_cells >= MAX_WORKBOOK_CELLS:
            truncated = True
            break
        cached_sheet = value_book[sheet.title] if sheet.title in value_book.sheetnames else None
        cells: list[SpreadsheetCell] = []
        text_parts.append(f"## Sheet: {sheet.title} ({sheet.sheet_state})")
        materialized_cells = sorted(getattr(sheet, "_cells", {}).values(), key=lambda cell: (cell.row, cell.column))
        for cell in materialized_cells:
            if total_cells >= MAX_WORKBOOK_CELLS:
                truncated = True
                break
            raw_value = cell.value
            if raw_value is None:
                continue
            formula = str(raw_value) if cell.data_type == "f" or (isinstance(raw_value, str) and raw_value.startswith("=")) else ""
            cached_value = cached_sheet[cell.coordinate].value if cached_sheet is not None else None
            display_value = cached_value if formula and cached_value is not None else raw_value
            cells.append(
                SpreadsheetCell(
                    coordinate=cell.coordinate,
                    value=display_value,
                    formula=formula,
                    cached_value=cached_value if formula else None,
                    data_type=str(cell.data_type or ""),
                    number_format=str(cell.number_format or ""),
                )
            )
            total_cells += 1
            formula_count += int(bool(formula))
            rendered = f"{cell.coordinate}: {display_value}"
            if formula:
                rendered += f" [formula: {formula}]"
                if cached_value is not None:
                    rendered += f" [cached: {cached_value}]"
            text_parts.append(rendered)
        hidden_rows = [index for index, dimension in sheet.row_dimensions.items() if dimension.hidden]
        hidden_columns = [index for index, dimension in sheet.column_dimensions.items() if dimension.hidden]
        sheets.append(
            SpreadsheetSheet(
                title=sheet.title,
                state=sheet.sheet_state,
                max_row=int(sheet.max_row or 0),
                max_column=int(sheet.max_column or 0),
                merged_ranges=[str(item) for item in sheet.merged_cells.ranges],
                hidden_rows=hidden_rows,
                hidden_columns=hidden_columns,
                cells=cells,
                truncated=truncated and total_cells >= MAX_WORKBOOK_CELLS,
            )
        )
    defined_names = workbook_defined_names(formula_book)
    external_links = workbook_external_links(formula_book)
    if truncated:
        warnings.append(ParseWarning(code="xlsx_structure_truncated", message="Workbook structure exceeded the safe parser resource boundary and was partially retained.", severity="warning"))
    workbook = SpreadsheetWorkbook(
        filename=filename,
        sheets=sheets,
        defined_names=defined_names,
        external_links=external_links,
        cell_count=total_cells,
        formula_count=formula_count,
        truncated=truncated,
    )
    capabilities = workbook_capabilities(workbook)
    return ParsedDocument(
        plain_text="\n".join(text_parts),
        outline=[OutlineItem(level=2, title=sheet.title, text=sheet.title) for sheet in sheets],
        warnings=warnings,
        capabilities=capabilities,
        workbooks=[workbook],
    )


def xlsx_archive_boundary_reason(content: bytes) -> str:
    try:
        with ZipFile(BytesIO(content)) as archive:
            entries = archive.infolist()
    except BadZipFile:
        return ""
    if len(entries) > MAX_XLSX_ARCHIVE_ENTRIES:
        return "Workbook archive contains too many entries to parse safely."
    uncompressed = sum(max(0, int(entry.file_size or 0)) for entry in entries)
    compressed = sum(max(0, int(entry.compress_size or 0)) for entry in entries)
    if uncompressed > MAX_XLSX_UNCOMPRESSED_BYTES:
        return "Workbook expanded content exceeds the safe parser resource boundary."
    if compressed > 0 and uncompressed / compressed > MAX_XLSX_COMPRESSION_RATIO:
        return "Workbook archive compression ratio exceeds the safe parser resource boundary."
    return ""


def workbook_defined_names(workbook: Any) -> list[dict[str, Any]]:
    result: list[dict[str, Any]] = []
    values = getattr(getattr(workbook, "defined_names", None), "values", None)
    names = list(values()) if callable(values) else list(getattr(getattr(workbook, "defined_names", None), "definedName", []) or [])
    for item in names:
        result.append(
            {
                "name": str(getattr(item, "name", "") or ""),
                "value": str(getattr(item, "attr_text", "") or ""),
                "hidden": bool(getattr(item, "hidden", False)),
                "local_sheet_id": getattr(item, "localSheetId", None),
            }
        )
    return result


def workbook_external_links(workbook: Any) -> list[str]:
    links: list[str] = []
    for index, item in enumerate(getattr(workbook, "_external_links", []) or [], start=1):
        target = str(getattr(getattr(item, "file_link", None), "Target", "") or "")
        links.append(target or f"external-link-{index}")
    return links


def workbook_capabilities(workbook: SpreadsheetWorkbook) -> list[MaterialCapability]:
    hidden_count = sum(len(sheet.hidden_rows) + len(sheet.hidden_columns) for sheet in workbook.sheets)
    cached_count = sum(1 for sheet in workbook.sheets for cell in sheet.cells if cell.formula and cell.cached_value is not None)
    return [
        MaterialCapability(id="text", status="partial" if workbook.truncated else "available", count=workbook.cell_count, detail="Cell values are represented with stable coordinates."),
        MaterialCapability(id="tables", status="unknown", detail="Workbook ranges are available; semantic table boundaries are not inferred."),
        MaterialCapability(id="sheet_structure", status="partial" if workbook.truncated else "available", count=len(workbook.sheets)),
        MaterialCapability(id="cell_coordinates", status="partial" if workbook.truncated else "available", count=workbook.cell_count),
        MaterialCapability(id="spreadsheet_formulas", status="partial" if workbook.truncated else "available", count=workbook.formula_count, detail="Raw formulas are preserved but never executed."),
        MaterialCapability(id="cached_formula_values", status="available" if cached_count else "unknown", count=cached_count, detail="Cached values are present only when stored by the source workbook."),
        MaterialCapability(id="defined_names", status="available", count=len(workbook.defined_names)),
        MaterialCapability(id="hidden_structure", status="available", count=hidden_count),
        MaterialCapability(id="external_links", status="available", count=len(workbook.external_links), detail="External links are identified but never accessed."),
    ]


def parse_document_basic(
    *,
    filename: str,
    content: bytes,
    content_type: str = "",
    reference_role: str = "material",
) -> ParsedDocument:
    suffix = Path(filename).suffix.lower()
    text = decode_content(content)
    warnings: list[ParseWarning] = []
    source_files = [SourceFile(filename=filename, content_type=content_type, size=len(content), role=reference_role)]

    if suffix in {".html", ".htm"} or "html" in content_type:
        parsed = parse_html_text(text)
    elif suffix in {".md", ".markdown"} or "markdown" in content_type:
        parsed = parse_markdown_text(text)
    elif suffix in {".txt"} or "plain" in content_type or not suffix:
        parsed = parse_plain_text(text)
    else:
        parsed = ParsedDocument(
            source_files=source_files,
            plain_text=normalize_text(text),
            warnings=[ParseWarning(code="unsupported_basic_parser", message=f"Unsupported file type for basic parser: {filename}", severity="warning")],
        )
        return replace(parsed, capabilities=generic_document_capabilities(parsed, filename=filename))

    parsed = replace(parsed, source_files=source_files, warnings=[*parsed.warnings, *warnings])
    return replace(parsed, capabilities=generic_document_capabilities(parsed, filename=filename))


def with_source_file(parsed: ParsedDocument, *, filename: str, content_type: str, size: int, reference_role: str) -> ParsedDocument:
    capabilities = parsed.capabilities or generic_document_capabilities(parsed, filename=filename)
    return replace(parsed, source_files=[SourceFile(filename=filename, content_type=content_type, size=size, role=reference_role)], capabilities=capabilities)


def generic_document_capabilities(parsed: ParsedDocument, *, filename: str) -> list[MaterialCapability]:
    suffix = Path(filename).suffix.lower()
    capabilities = [
        MaterialCapability(id="text", status="available" if parsed.plain_text else "unavailable", count=len(parsed.plain_text)),
        MaterialCapability(id="tables", status="available" if parsed.tables else "unknown", count=len(parsed.tables)),
    ]
    if suffix in {".xlsx", ".xls"}:
        capabilities.extend(
            [
                MaterialCapability(id="sheet_structure", status="partial", detail="Generic extraction may preserve sheet labels but not complete workbook structure."),
                MaterialCapability(id="cell_coordinates", status="unavailable", detail="The generic parser does not preserve stable cell coordinates."),
                MaterialCapability(id="spreadsheet_formulas", status="unavailable", detail="The generic parser exposes display text, not raw workbook formulas."),
                MaterialCapability(id="cached_formula_values", status="unknown"),
                MaterialCapability(id="defined_names", status="unavailable"),
                MaterialCapability(id="hidden_structure", status="unavailable"),
                MaterialCapability(id="external_links", status="unknown"),
            ]
        )
    return capabilities


def parse_html_text(text: str) -> ParsedDocument:
    parser = _TextExtractor()
    try:
        parser.feed(text)
    except Exception:
        parser.close()
    outline = [OutlineItem(level=level, title=title, text=title) for level, title in parser.headings]
    plain_text = normalize_text(" ".join(parser.chunks))
    return ParsedDocument(
        plain_text=plain_text,
        outline=outline,
        links=parser.links,
        style_hints=infer_style_hints(text, kind="html"),
    )


def parse_markdown_text(text: str) -> ParsedDocument:
    headings: list[OutlineItem] = []
    lines = text.splitlines()
    for line in lines:
        match = re.match(r"^(#{1,6})\s+(.+)$", line.strip())
        if match:
            level = len(match.group(1))
            title = normalize_text(match.group(2))
            if title:
                headings.append(OutlineItem(level=level, title=title, text=title))
    plain_text = normalize_text(text)
    return ParsedDocument(
        plain_text=plain_text,
        outline=headings,
        style_hints=infer_style_hints(text, kind="markdown"),
    )


def parse_plain_text(text: str) -> ParsedDocument:
    plain_text = normalize_text(text)
    headings = infer_text_outline(text)
    return ParsedDocument(
        plain_text=plain_text,
        outline=headings,
        style_hints=infer_style_hints(text, kind="text"),
    )


def infer_style_hints(text: str, *, kind: str) -> list[StyleHint]:
    hints: list[StyleHint] = []
    sample = text[:2000]
    if any(token in sample for token in ("#", "标题", "Heading", "title")):
        hints.append(StyleHint(kind=f"{kind}:heading", value="structured_heading", confidence=0.5))
    if any(token in sample for token in ("table", "|", "表格")):
        hints.append(StyleHint(kind=f"{kind}:table", value="tabular", confidence=0.4))
    if any(token in sample for token in ("dark", "#111", "#000")):
        hints.append(StyleHint(kind=f"{kind}:palette", value="dark", confidence=0.4))
    if any(token in sample for token in ("blue", "#0f766e", "#2563eb")):
        hints.append(StyleHint(kind=f"{kind}:palette", value="blue", confidence=0.4))
    return hints


def infer_text_outline(text: str) -> list[OutlineItem]:
    outline: list[OutlineItem] = []
    for line in text.splitlines():
        cleaned = normalize_text(line)
        if not cleaned:
            continue
        if len(cleaned) <= 80 and cleaned.endswith((":", "。")):
            outline.append(OutlineItem(level=2, title=cleaned.rstrip("：:。"), text=cleaned))
    return outline[:12]


def decode_content(content: bytes) -> str:
    for encoding in ("utf-8", "utf-8-sig", "gb18030", "latin-1"):
        try:
            return content.decode(encoding)
        except UnicodeDecodeError:
            continue
    return content.decode("utf-8", errors="ignore")


def normalize_text(value: Any) -> str:
    return re.sub(r"\s+", " ", str(value or "").strip())


def normalize_parser_mode(value: str) -> str:
    normalized = str(value or "").strip().lower()
    if normalized == PARSER_BASIC:
        return PARSER_BASIC
    return PARSER_MARKITDOWN


def blocking_parse_failure_reason(parsed: ParsedDocument, *, filename: str = "", content_type: str = "") -> str:
    resource_warning = next((warning.message for warning in parsed.warnings if warning.code == "xlsx_resource_boundary"), "")
    if resource_warning:
        return resource_warning
    if not requires_enhanced_parser(filename=filename, content_type=content_type):
        return ""
    codes = {warning.code for warning in parsed.warnings}
    if "unsupported_basic_parser" not in codes:
        return ""
    if not ({"markitdown_unavailable_or_failed", "enhanced_parser_disabled"} & codes):
        return ""
    text = str(parsed.plain_text or "")
    if looks_like_container_or_binary_noise(text):
        return (
            f"Failed to parse {filename or 'uploaded material'} into readable text. "
            "The enhanced parser was unavailable or failed, and local fallback returned container/binary content."
        )
    if len(normalize_text(text)) < 80:
        return (
            f"Failed to parse {filename or 'uploaded material'} into enough readable text. "
            "The enhanced parser was unavailable or failed, and local fallback produced too little content."
        )
    return ""


def requires_enhanced_parser(*, filename: str = "", content_type: str = "") -> bool:
    suffix = Path(filename).suffix.lower()
    if suffix in BASIC_SUFFIXES or any(token in content_type for token in ("html", "markdown", "plain")):
        return False
    return suffix in MARKITDOWN_SUFFIXES or any(
        token in content_type
        for token in (
            "pdf",
            "word",
            "presentation",
            "spreadsheet",
            "excel",
            "image/",
        )
    )


def looks_like_container_or_binary_noise(text: str) -> bool:
    sample = str(text or "")[:12000]
    if not sample:
        return True
    lowered = sample.lower()
    container_markers = (
        "pk\x03\x04",
        "[content_types].xml",
        "word/document.xml",
        "ppt/slides/",
        "xl/workbook.xml",
        "%pdf-",
        "endobj",
        "xref",
    )
    if any(marker in lowered for marker in container_markers):
        return True
    if len(sample) < 120:
        return False
    control_chars = sum(1 for char in sample if ord(char) < 32 and char not in "\n\r\t")
    replacement_chars = sample.count("\ufffd")
    odd_chars = sum(1 for char in sample if 127 <= ord(char) <= 159)
    return (control_chars + replacement_chars + odd_chars) / max(1, len(sample)) > 0.02


def parse_with_markitdown(filename: str, content: bytes) -> ParsedDocument | None:
    if MarkItDown is None:
        return None
    try:  # pragma: no cover - optional dependency
        return parse_with_retry(lambda: parse_with_markitdown_once(filename, content), parser_name="markitdown")
    except Exception as exc:  # pragma: no cover - optional dependency
        return ParsedDocument(
            plain_text="",
            warnings=[ParseWarning(code="markitdown_failed", message=str(exc), severity="warning")],
        )


def parse_with_markitdown_once(filename: str, content: bytes) -> ParsedDocument:
    suffix = Path(filename).suffix
    parser = MarkItDown()
    with tempfile.NamedTemporaryFile(suffix=suffix) as handle:
        handle.write(content)
        handle.flush()
        result = parser.convert(handle.name)
    text = normalize_text(getattr(result, "text_content", "") or "")
    return ParsedDocument(
        plain_text=text,
        warnings=[ParseWarning(code="markitdown_used", message="Parsed via MarkItDown.", severity="info")],
    )


def parse_with_retry(parse_fn, *, parser_name: str) -> ParsedDocument:
    last_exc: Exception | None = None
    for attempt in range(1, PARSER_RETRY_ATTEMPTS + 1):
        try:
            parsed = parse_fn()
            if attempt > 1:
                retry_warning = ParseWarning(code=f"{parser_name}_retry_succeeded", message=f"{parser_name} parser succeeded on retry {attempt}.", severity="info")
                return replace(parsed, warnings=[*parsed.warnings, retry_warning])
            return parsed
        except Exception as exc:
            last_exc = exc
    raise last_exc or RuntimeError(f"{parser_name} parser failed.")
